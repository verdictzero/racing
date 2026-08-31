#!/usr/bin/env python3
"""
Regenerate packages/core/src/__fixtures__/foreign-workbook.xlsx.

Deliberately uses openpyxl rather than this repo's own writer: the fixture exists to prove the
importer can read a workbook written by something else — shared strings, deflate compression, and
the parts Excel includes that ours does not. Run with `pip install openpyxl` available.
"""
import pathlib
import openpyxl

OUT = pathlib.Path(__file__).resolve().parent.parent / 'packages/core/src/__fixtures__/foreign-workbook.xlsx'

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'RACI'
ws.append(['Cyber Directorate RACI'])      # preamble the header scan must look past
ws.append([])
ws.append(['Portfolio', 'Program', 'Project', 'Task', 'Org unit', 'R&D', 'Legal Ops'])
ws.append(['Alpha', '', '', '', 'DIRECTORATE C', 'A', 'R'])
ws.append(['Alpha', 'Beta', '', '', 'Cyber', 'ar', 'C'])
ws.append(['Alpha', '', 'Orphan', '', '', 'A', 'R'])   # a gap: must be skipped and counted

ent = wb.create_sheet('Entities')
ent.append(['Entity', 'Kind', 'Short', 'Lead', 'Description'])
ent.append(['Cyber Review Board', 'board', 'CRB', 'A. Person', 'Reviews & approves'])

doc = wb.create_sheet('Document')
doc.append(['Kind', 'Name', 'Status', 'Signed', 'Customer', 'Priority', 'Budget', 'Tags', 'Description'])
doc.append(['Chart', 'Cyber Response', 'Draft', '', 'J3', 'normal', '$1m', 'cyber, ir', 'From openpyxl'])

wb.save(OUT)
print(f'wrote {OUT}')
